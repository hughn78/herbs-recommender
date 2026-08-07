"""Phase 2 — corpus analysis.

Reads the Herbs of Gold source corpus and the prior extraction pipeline's
outputs, cross-compares the four source formats, and emits machine-readable
audit reports under data/reports/:

  source_inventory.json     every source file, size, sha256, format
  corpus_audit.json         the headline coverage/quality audit
  cross_source.json         per-product presence across PDF/DOCX/XLSX/MD
  conflicts.json            cross-source conflicts + extraction issues

Never writes to the corpus. Never writes to the application database
(that is the Phase 5 ingestion pipeline's job).
"""

from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from . import corpus


def normalise_name(name: str) -> str:
    n = name.lower().replace("’", "'").replace("‘", "'")
    n = re.sub(r"[^a-z0-9'&+ ]+", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def markdown_product_headings(md_path: Path, known_names: set[str]) -> dict[str, int]:
    """Count markdown headings that match a known product name."""
    counts: dict[str, int] = {}
    if not md_path.exists():
        return counts
    for line in md_path.read_text(encoding="utf-8", errors="replace").splitlines():
        m = re.match(r"^#{1,3}\s+(.*\S)\s*$", line)
        if not m:
            continue
        n = normalise_name(m.group(1))
        if n in known_names:
            counts[n] = counts.get(n, 0) + 1
    return counts


def analyse() -> dict:
    products = corpus.load_json(corpus.KB_PRODUCTS_JSON)
    issues = corpus.load_json(corpus.KB_ISSUES_JSON)
    docx_names = corpus.load_json(corpus.KB_DOCX_NAMES)
    excel_toc = corpus.load_json(corpus.KB_EXCEL_TOC)  # {normalised_name: page}
    pdf_index = corpus.load_json(corpus.KB_PDF_PAGE_INDEX)

    # ---- per-product cross-source presence -------------------------------
    by_norm = {normalise_name(p["product_name"]): p for p in products}
    known = set(by_norm)

    if isinstance(docx_names, dict):
        docx_set = {normalise_name(k) for k in docx_names}
    else:
        docx_set = {normalise_name(n) for n in docx_names}
    excel_set = {normalise_name(k) for k in excel_toc}
    # pdf_page_index: {page: product_name} or list — normalise to name set
    if isinstance(pdf_index, dict):
        pdf_names = {normalise_name(str(v)) for v in pdf_index.values()}
    else:
        pdf_names = {normalise_name(str(v)) for v in pdf_index}
    pdf_set = {n for n in pdf_names if n in known}

    md_counts = markdown_product_headings(
        corpus.CORPUS_ROOT / "markdown" / "herbsofgold_technical_manual.md", known
    )

    cross_source = []
    for norm, p in sorted(by_norm.items()):
        cross_source.append(
            {
                "product_id": p["product_id"],
                "product_name": p["product_name"],
                "in_docx": norm in docx_set,
                "in_xlsx_toc": norm in excel_set,
                "in_pdf_index": norm in pdf_set,
                "in_markdown": norm in md_counts,
            }
        )

    toc_only = sorted(excel_set - docx_set)

    # ---- coverage / quality ----------------------------------------------
    def has_ingredients(p):
        return bool(p.get("ingredients"))

    def has_indications(p):
        return any((i.get("text") or "").strip() for i in p.get("indications", []))

    def has_cautions(p):
        return any((c.get("text") or "").strip() for c in p.get("cautions", []))

    def has_dose(p):
        d = p.get("directions") or {}
        return bool((d.get("adult_dose") or "").strip() or (d.get("child_dose") or "").strip())

    conf_counter = Counter()
    # Product-level ExtractionConfidence lives in the CSV export (the JSON
    # only carries per-ingredient confidence).
    csv_path = corpus.KB_ROOT / "output" / "herbs_of_gold_products.csv"
    if csv_path.exists():
        import csv as _csv

        with csv_path.open(encoding="utf-8") as fh:
            for row in _csv.DictReader(fh):
                conf_counter[row.get("ExtractionConfidence") or "unknown"] += 1
    else:
        for p in products:
            rev = p.get("review") or {}
            c = rev.get("extraction_confidence") or rev.get("confidence") or "unknown"
            conf_counter[str(c)] += 1

    name_counter = Counter(normalise_name(p["product_name"]) for p in products)
    duplicate_names = sorted(n for n, c in name_counter.items() if c > 1)

    audit = {
        "total_products": len(products),
        "duplicate_product_names": duplicate_names,
        "toc_only_entries_not_in_catalogue": toc_only,
        "products_missing_ingredients": [
            p["product_id"] for p in products if not has_ingredients(p)
        ],
        "products_missing_indications": [
            p["product_id"] for p in products if not has_indications(p)
        ],
        "products_missing_warnings": [
            p["product_id"] for p in products if not has_cautions(p)
        ],
        "products_missing_dosage": [p["product_id"] for p in products if not has_dose(p)],
        "products_missing_dosage_form": [
            p["product_id"] for p in products if not (p.get("dosage_form") or "").strip()
        ],
        "products_missing_austl": [
            p["product_id"] for p in products if not (p.get("austl") or "").strip()
        ],
        "extraction_confidence_distribution": dict(conf_counter),
        "unique_ingredient_names": len(
            {
                (i.get("ingredient_name") or "").strip().lower()
                for p in products
                for i in p.get("ingredients", [])
                if (i.get("ingredient_name") or "").strip()
            }
        ),
        "totals": {
            "clinical_use_tags": sum(
                len((p.get("clinical_tags") or {}).get("clinical_use_tags") or [])
                for p in products
            ),
            "avoid_if_tags": sum(
                len((p.get("clinical_tags") or {}).get("avoid_if_tags") or [])
                for p in products
            ),
            "extraction_issues_logged": len(issues),
        },
    }

    # ---- conflicts ---------------------------------------------------------
    issue_type_counter = Counter(i.get("IssueType", "?") for i in issues)
    conflicts = {
        "extraction_issues_by_type": dict(issue_type_counter),
        "extraction_issues": issues,
        "word_excel_inconsistencies": _read_validation_sheet("Word_Excel_Inconsistencies"),
        "possible_errors": _read_validation_sheet("Possible_Errors"),
    }

    corpus.write_report("source_inventory.json", corpus.inventory_sources())
    corpus.write_report("corpus_audit.json", audit)
    corpus.write_report("cross_source.json", cross_source)
    corpus.write_report("conflicts.json", conflicts)
    return audit


def _read_validation_sheet(sheet_name: str) -> list[dict]:
    """Read one sheet of the prior pipeline's validation workbook."""
    if not corpus.KB_VALIDATION_XLSX.exists():
        return []
    wb = load_workbook(corpus.KB_VALIDATION_XLSX, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    out = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        out.append({header[i]: (str(v) if v is not None else "") for i, v in enumerate(r)})
    return out


if __name__ == "__main__":
    result = analyse()
    print(f"products={result['total_products']} "
          f"missing_ingredients={len(result['products_missing_ingredients'])} "
          f"missing_warnings={len(result['products_missing_warnings'])} "
          f"missing_dosage={len(result['products_missing_dosage'])} "
          f"confidence={result['extraction_confidence_distribution']}",
          file=sys.stderr)
